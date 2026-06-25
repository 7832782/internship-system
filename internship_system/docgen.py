"""文档生成工具 —— 实习报告 Word + 登记表/监管信息 Excel"""
import io, re, os
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = BASE_DIR / 'templates'

TEMPLATE1 = str(TEMPLATE_DIR / 'template1.docx')
TEMPLATE2 = str(TEMPLATE_DIR / 'template2.xlsx')
TEMPLATE3 = str(TEMPLATE_DIR / 'template3.xlsx')


# ─── 辅助函数 ───────────────────────────────

def get_display_width(text):
    """中文字符占2，半角占1"""
    return sum(2 if ord(c) > 127 else 1 for c in text)


def pad_to_width_center(text, target_width):
    """将字符串居中填充到指定显示宽度"""
    w = get_display_width(text)
    if w >= target_width:
        return text
    p = target_width - w
    left = p // 2
    return ' ' * left + text + ' ' * (p - left)


def format_date(date_str):
    """日期格式化为 2026/04/30"""
    if not date_str:
        return ''
    try:
        if '年' in date_str:
            m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
            if m:
                y, mo, d = m.groups()
                return f"{y}/{mo.zfill(2)}/{d.zfill(2)}"
        if '-' in date_str:
            parts = date_str.split('-')
            return f"{parts[0]}/{parts[1].zfill(2)}/{parts[2].zfill(2)}"
        if '/' in date_str:
            parts = date_str.split('/')
            return f"{parts[0]}/{parts[1].zfill(2)}/{parts[2].zfill(2)}"
    except Exception:
        pass
    return date_str


def parse_date_ymd(date_str):
    """将日期解析为(年,月,日)"""
    if not date_str:
        return '', '', ''
    try:
        if '-' in date_str:
            parts = date_str.split('-')
            return parts[0], parts[1], parts[2]
        if '/' in date_str:
            parts = date_str.split('/')
            return parts[0], parts[1], parts[2]
        if '年' in date_str:
            m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
            if m:
                return m.groups()
    except Exception:
        pass
    return '', '', ''


def val(v):
    """空值转为空字符串。

    注意：不能用 `if v` 判断，否则数值 0 / 0.0 / 布尔 False 会被错误清空。
    只把 None、空字符串、空值归一为 ''，保留合法的 0。
    """
    if v is None:
        return ''
    if isinstance(v, str) and v == '':
        return ''
    return v


# ─── 生成实习报告 Word ───────────────────────

def generate_report(data):
    """根据模板生成实习报告 Word 文档"""
    from docxtpl import DocxTemplate

    tpl = DocxTemplate(TEMPLATE1)

    college = data.get('college', '')
    major = data.get('major', '')
    # 以下原对齐填充逻辑已注释（模板内已处理对齐）
    cw = get_display_width(college)
    mw = get_display_width(major)
    if cw + mw < 32:
        r = 32 - cw - mw
        h = r // 2
        college = pad_to_width_center(college, cw + h)
        major = pad_to_width_center(major, mw + r - h)

    name = data.get('name', '')
    sid = data.get('studentId', '')
    nw = get_display_width(name)
    sw = get_display_width(sid)
    if nw + sw < 38:
        r = 38 - nw - sw
        h = r // 2
        name = pad_to_width_center(name, nw + h)
        sid = pad_to_width_center(sid, sw + r - h)

    teacher = data.get('teacher', '')
    tw = get_display_width(teacher)
    if tw < 36:
        teacher = pad_to_width_center(teacher, 36)

    company = data.get('company', '')
    cow = get_display_width(company)
    if cow < 40:
        company = pad_to_width_center(company, 40)

    start_time = format_date(data.get('startDate', ''))
    end_time = format_date(data.get('endDate', ''))
    submit_time = format_date(data.get('submitDate', ''))

    total_days = data.get('totalDays', '0')
    try:
        td = ''.join(filter(str.isdigit, str(total_days))) or '0'
    except Exception:
        td = '0'
    internship_time = td

    sw_ = get_display_width(start_time)
    ew = get_display_width(end_time)
    tw = get_display_width(internship_time)
    if sw_ + ew + tw < 28:
        r = 28 - sw_ - ew - tw
        t = r // 3
        start_time = pad_to_width_center(start_time, sw_ + t)
        end_time = pad_to_width_center(end_time, ew + t)
        internship_time = pad_to_width_center(internship_time, tw + r - 2 * t)

    sub_w = get_display_width(submit_time)
    if sub_w < 32:
        submit_time = pad_to_width_center(submit_time, 32)

    context = {
        'college': college,
        'major': major,
        'name': name,
        'student_id': sid,
        'guide_teacher': teacher,
        'internship_unit': company,
        'start_time': start_time,
        'end_time': end_time,
        'internship_time': internship_time,
        'submit_time': submit_time,
        'internship_content': data.get('internshipContent', ''),
        'internship_lessons': data.get('experience', ''),
        'suggestions': data.get('suggestions', ''),
    }

    tpl.render(context)
    output = io.BytesIO()
    tpl.docx.save(output)
    output.seek(0)
    return output


# ─── 生成实习登记表 Excel ──────────────────

def generate_registration_excel(data):
    """根据模板生成实习登记表 Excel"""
    from openpyxl import load_workbook

    wb = load_workbook(TEMPLATE2)
    ws = wb.active

    start_y, start_m, start_d = parse_date_ymd(data.get('startDate', ''))
    end_y, end_m, end_d = parse_date_ymd(data.get('endDate', ''))

    total_days = data.get('totalDays', '0')
    try:
        td = ''.join(filter(str.isdigit, str(total_days))) or '0'
    except Exception:
        td = '0'

    replacements = {
        '{{student_id}}': val(data.get('studentId')),
        '{{name}}': val(data.get('name')),
        '{{gender}}': val(data.get('gender')),
        '{{college}}': val(data.get('college')),
        '{{major}}': val(data.get('major')),
        '{{internship_unit}}': val(data.get('company')),
        '{{start_time_y}}': start_y or '无',
        '{{start_time_m}}': start_m or '无',
        '{{start_time_d}}': start_d or '无',
        '{{end_time_y}}': end_y or '无',
        '{{end_time_m}}': end_m or '无',
        '{{end_time_d}}': end_d or '无',
        '{{internship_time}}': str(td) if td else '无',
        '{{self_evaluation}}': val(data.get('selfEvaluation')),
    }

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, value in replacements.items():
                    if key in cell.value:
                        cell.value = cell.value.replace(key, value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─── 生成实习监管信息 Excel ────────────────

def generate_supervision_excel(data):
    """使用 template3.xlsx 模板生成实习监管信息 Excel"""
    from openpyxl import load_workbook

    wb = load_workbook(TEMPLATE3)
    ws = wb.active
    # 清除模板中多余的空行，从第 2 行开始删
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    row_data = [
        val(data.get('studentId')), val(data.get('name')),
        val(data.get('enrollmentYear')), val(data.get('college')),
        val(data.get('classNumber')), val(data.get('courseName')),
        val(data.get('courseCode')), val(data.get('credits')),
        val(data.get('internshipType')), val(data.get('internshipOrgForm')),
        val(data.get('internshipMethod')), val(data.get('academicYear')),
        val(data.get('teacher')), val(data.get('company')),
        val(data.get('companyCreditCode')), val(data.get('internshipAreaCode')),
        val(data.get('internshipAddress')), val(data.get('hasLiabilityInsurance')),
        val(data.get('hasAccidentInsurance')), val(data.get('hasSafetyTraining')),
        val(data.get('hasTripartiteAgreement')), val(data.get('companyAcquisitionMethod')),
        val(data.get('hasEmergencyPlan')), val(data.get('isSchoolBase')),
        val(data.get('isDigitalBase')), val(data.get('isOverseasBase')),
        val(data.get('startDate')), val(data.get('endDate')),
        val(data.get('totalDays')) if data.get('totalDays') is not None else '',
        val(data.get('internshipPosition')),
        val(data.get('internshipSalary')), val(data.get('companySupervisor')),
    ]
    ws.append(row_data)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_supervision_batch_excel(rows):
    """生成批量实习监管信息汇总 Excel"""
    from openpyxl import load_workbook

    wb = load_workbook(TEMPLATE3)
    ws = wb.active
    # 清除模板中多余的空行
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_data in rows:
        ws.append([
            val(row_data.get('studentId')), val(row_data.get('name')),
            val(row_data.get('enrollmentYear')), val(row_data.get('college')),
            val(row_data.get('classNumber')), val(row_data.get('courseName')),
            val(row_data.get('courseCode')), val(row_data.get('credits')),
            val(row_data.get('internshipType')), val(row_data.get('internshipOrgForm')),
            val(row_data.get('internshipMethod')), val(row_data.get('academicYear')),
            val(row_data.get('teacher')), val(row_data.get('company')),
            val(row_data.get('companyCreditCode')), val(row_data.get('internshipAreaCode')),
            val(row_data.get('internshipAddress')), val(row_data.get('hasLiabilityInsurance')),
            val(row_data.get('hasAccidentInsurance')), val(row_data.get('hasSafetyTraining')),
            val(row_data.get('hasTripartiteAgreement')), val(row_data.get('companyAcquisitionMethod')),
            val(row_data.get('hasEmergencyPlan')), val(row_data.get('isSchoolBase')),
            val(row_data.get('isDigitalBase')), val(row_data.get('isOverseasBase')),
            val(row_data.get('startDate')), val(row_data.get('endDate')),
            val(row_data.get('totalDays')) if row_data.get('totalDays') is not None else '',
            val(row_data.get('internshipPosition')),
            val(row_data.get('internshipSalary')), val(row_data.get('companySupervisor')),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
