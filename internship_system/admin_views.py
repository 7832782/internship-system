"""管理后台 API 视图 - 通用列表/CRUD"""
import math, io, zipfile
from datetime import datetime
from django.db.models import Q
from django.db import IntegrityError
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.models import User
from internship.models import Internship
from internship.serializers import InternshipSerializer
from registration.models import Registration
from registration.serializers import RegistrationSerializer
from supervision.models import Supervision
from supervision.serializers import SupervisionSerializer
from accounts.serializers import UserSerializer
from internship_system.docgen import generate_report, generate_registration_excel, generate_supervision_batch_excel
from internship_system.models import SystemSetting


def admin_required(view_func):
    """装饰器：检查管理员权限"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_admin:
            return Response({'success': False, 'message': '权限不足'},
                            status=status.HTTP_403_FORBIDDEN)
        return view_func(request, *args, **kwargs)
    return wrapper


def paginate(queryset, request, search_fields=None, filter_map=None):
    """通用分页 + 搜索 + 过滤"""
    page = int(request.query_params.get('page', 1))
    per_page = int(request.query_params.get('per_page', 50))
    keyword = request.query_params.get('keyword', '').strip()
    if per_page < 1:
        per_page = 50
    if page < 1:
        page = 1

    # 搜索
    if keyword and search_fields:
        q = Q()
        for field in search_fields:
            q |= Q(**{field + '__icontains': keyword})
        queryset = queryset.filter(q)

    # 过滤
    if filter_map:
        for param, field in filter_map.items():
            val = request.query_params.get(param, '').strip()
            if val:
                queryset = queryset.filter(**{field: val})

    total = queryset.count()
    total_pages = max(1, math.ceil(total / per_page))
    offset = (page - 1) * per_page
    items = queryset[offset:offset + per_page]

    return {
        'page': page,
        'total_pages': total_pages,
        'total': total,
        'data': items,
        '_filtered_qs': queryset if request.query_params.get('all_ids') else None,
    }


def flatten_student(item_dict, item_obj):
    """为字典添加关联学生的平铺字段（studentId, batch, classNumber, major）"""
    try:
        student = item_obj.student
        item_dict['studentId'] = student.student_id
        item_dict['batch'] = student.batch
        item_dict['classNumber'] = student.class_number
        item_dict.setdefault('major', student.major)
    except Exception:
        item_dict['studentId'] = getattr(item_obj, 'student_id', '')
        item_dict['batch'] = '无'
        item_dict['classNumber'] = ''
        item_dict.setdefault('major', '')
    return item_dict


# ─── 实习报告管理 ───────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_internship_list(request):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    if request.method == 'GET':
        qs = Internship.objects.select_related('student').all()
        filter_map = {
            'college': 'college__icontains',
            'major': 'major__icontains',
            'company': 'company__icontains',
            'batch': 'student__batch',
        }
        result = paginate(qs, request, search_fields=['name', 'company', 'student__student_id'], filter_map=filter_map)
        raw_items = result['data']
        result['data'] = InternshipSerializer(raw_items, many=True).data
        result['data'] = [flatten_student(d, obj) for d, obj in zip(result['data'], raw_items)]
        if request.query_params.get('all_ids'):
            result['ids'] = list(result['_filtered_qs'].values_list('student_id', flat=True))
        result.pop('_filtered_qs', None)
        return Response(result)
    student_id = request.data.get('student_id')
    if not student_id:
        return Response({'success': False, 'message': '学号不能为空'}, status=400)
    try:
        student = User.objects.get(student_id=student_id)
    except User.DoesNotExist:
        # 学号不存在时自动创建用户（方便老师临时补充学生数据）
        student = User.objects.create_user(
            username=student_id,
            student_id=student_id,
            password=student_id,
            name=request.data.get('name', ''),
            college=request.data.get('college', ''),
            major=request.data.get('major', ''),
            batch=request.data.get('batch', '无'),
            role='student',
        )
    # 给数据库 NOT NULL 字段补默认值（老师可能只填了学号）
    import copy
    data = copy.deepcopy(dict(request.data))
    defaults = {
        'name': '',
        'college': '',
        'major': '',
        'teacher': '',
        'company': '',
    }
    for k, v in defaults.items():
        if k not in data or data[k] in (None, ''):
            data[k] = v
    for k in ['start_date', 'end_date', 'total_days']:
        if k not in data or data.get(k) in (None, ''):
            data[k] = None
    serializer = InternshipSerializer(data=data)
    if serializer.is_valid():
        try:
            serializer.save(student=student)
        except IntegrityError:
            return Response({'success': False, 'message': '该学生已有实习报告记录'}, status=400)
        return Response({'success': True, 'data': serializer.data}, status=201)
    return Response({'success': False, 'message': serializer.errors}, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_internship_detail(request, student_id):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    try:
        obj = Internship.objects.get(student_id=student_id)
    except Internship.DoesNotExist:
        return Response({'success': False, 'message': '记录不存在'}, status=404)
    if request.method == 'GET':
        return Response(InternshipSerializer(obj).data)
    elif request.method == 'PUT':
        data = request.data.copy()
        for k in ['start_date', 'end_date', 'total_days']:
            if k in data and data[k] in (None, ''):
                data[k] = None
        student = None
        sid = data.pop('student_id', None)
        if sid:
            try:
                student = User.objects.get(student_id=sid)
            except User.DoesNotExist:
                return Response({'success': False, 'message': '学生不存在'}, status=404)
        s = InternshipSerializer(obj, data=data, partial=True)
        if s.is_valid():
            kwargs = {}
            if student:
                kwargs['student'] = student
            try:
                s.save(**kwargs)
            except IntegrityError:
                return Response({'success': False, 'message': '该学生已有实习报告记录'}, status=400)
            return Response({'success': True, 'data': s.data})
        return Response({'success': False, 'message': s.errors}, status=400)
    elif request.method == 'DELETE':
        obj.delete()
        return Response({'success': True, 'message': '删除成功'})


# ─── 实习登记管理 ───────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_registration_list(request):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    if request.method == 'GET':
        qs = Registration.objects.select_related('student').all()
        filter_map = {
            'college': 'college__icontains',
            'major': 'major__icontains',
            'company': 'company__icontains',
            'batch': 'student__batch',
        }
        result = paginate(qs, request, search_fields=['name', 'company', 'student__student_id'], filter_map=filter_map)
        raw_items = result['data']
        result['data'] = RegistrationSerializer(raw_items, many=True).data
        result['data'] = [flatten_student(d, obj) for d, obj in zip(result['data'], raw_items)]
        if request.query_params.get('all_ids'):
            result['ids'] = list(result['_filtered_qs'].values_list('student_id', flat=True))
        result.pop('_filtered_qs', None)
        return Response(result)
    student_id = request.data.get('student_id')
    if not student_id:
        return Response({'success': False, 'message': '学号不能为空'}, status=400)
    try:
        student = User.objects.get(student_id=student_id)
    except User.DoesNotExist:
        # 学号不存在时自动创建用户（方便老师临时补充学生数据）
        student = User.objects.create_user(
            username=student_id,
            student_id=student_id,
            password=student_id,
            name=request.data.get('name', ''),
            college=request.data.get('college', ''),
            major=request.data.get('major', ''),
            batch=request.data.get('batch', '无'),
            role='student',
        )
    import copy
    data = copy.deepcopy(dict(request.data))
    for k in ['name', 'gender', 'college', 'major', 'company']:
        if k not in data or data[k] in (None, ''):
            data[k] = ''
    for k in ['start_date', 'end_date', 'total_days']:
        if k not in data or data.get(k) in (None, ''):
            data[k] = None
    serializer = RegistrationSerializer(data=data)
    if serializer.is_valid():
        try:
            serializer.save(student=student)
        except IntegrityError:
            return Response({'success': False, 'message': '该学生已有登记表记录'}, status=400)
        return Response({'success': True, 'data': serializer.data}, status=201)
    return Response({'success': False, 'message': serializer.errors}, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_registration_detail(request, student_id):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    try:
        obj = Registration.objects.get(student_id=student_id)
    except Registration.DoesNotExist:
        return Response({'success': False, 'message': '记录不存在'}, status=404)
    if request.method == 'GET':
        return Response(RegistrationSerializer(obj).data)
    elif request.method == 'PUT':
        data = request.data.copy()
        for k in ['start_date', 'end_date', 'total_days']:
            if k in data and data[k] in (None, ''):
                data[k] = None
        student = None
        sid = data.pop('student_id', None)
        if sid:
            try:
                student = User.objects.get(student_id=sid)
            except User.DoesNotExist:
                return Response({'success': False, 'message': '学生不存在'}, status=404)
        s = RegistrationSerializer(obj, data=data, partial=True)
        if s.is_valid():
            kwargs = {}
            if student:
                kwargs['student'] = student
            try:
                s.save(**kwargs)
            except IntegrityError:
                return Response({'success': False, 'message': '该学生已有登记表记录'}, status=400)
            return Response({'success': True, 'data': s.data})
        return Response({'success': False, 'message': s.errors}, status=400)
    elif request.method == 'DELETE':
        obj.delete()
        return Response({'success': True, 'message': '删除成功'})


# ─── 实习监管管理 ───────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_supervision_list(request):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    if request.method == 'GET':
        qs = Supervision.objects.select_related('student').all()
        filter_map = {
            'college': 'college__icontains',
            'company': 'company__icontains',
            'batch': 'student__batch',
        }
        result = paginate(qs, request, search_fields=['name', 'company', 'student__student_id'], filter_map=filter_map)
        raw_items = result['data']
        result['data'] = SupervisionSerializer(raw_items, many=True).data
        result['data'] = [flatten_student(d, obj) for d, obj in zip(result['data'], raw_items)]
        if request.query_params.get('all_ids'):
            result['ids'] = list(result['_filtered_qs'].values_list('student_id', flat=True))
        result.pop('_filtered_qs', None)
        return Response(result)
    student_id = request.data.get('student_id')
    if not student_id:
        return Response({'success': False, 'message': '学号不能为空'}, status=400)
    try:
        student = User.objects.get(student_id=student_id)
    except User.DoesNotExist:
        # 学号不存在时自动创建用户（方便老师临时补充学生数据）
        student = User.objects.create_user(
            username=student_id,
            student_id=student_id,
            password=student_id,
            name=request.data.get('name', ''),
            college=request.data.get('college', ''),
            major=request.data.get('major', ''),
            batch=request.data.get('batch', '无'),
            role='student',
        )
    import copy
    data = copy.deepcopy(dict(request.data))
    for k in ['name', 'college', 'company', 'internship_address']:
        if k not in data or data[k] in (None, ''):
            data[k] = ''
    for k in ['start_date', 'end_date', 'total_days']:
        if k not in data or data.get(k) in (None, ''):
            data[k] = None
    serializer = SupervisionSerializer(data=data)
    if serializer.is_valid():
        try:
            serializer.save(student=student)
        except IntegrityError:
            return Response({'success': False, 'message': '该学生已有监管信息记录'}, status=400)
        return Response({'success': True, 'data': serializer.data}, status=201)
    return Response({'success': False, 'message': serializer.errors}, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_supervision_detail(request, student_id):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    try:
        obj = Supervision.objects.get(student_id=student_id)
    except Supervision.DoesNotExist:
        return Response({'success': False, 'message': '记录不存在'}, status=404)
    if request.method == 'GET':
        return Response(SupervisionSerializer(obj).data)
    elif request.method == 'PUT':
        data = request.data.copy()
        for k in ['start_date', 'end_date', 'total_days']:
            if k in data and data[k] in (None, ''):
                data[k] = None
        student = None
        sid = data.pop('student_id', None)
        if sid:
            try:
                student = User.objects.get(student_id=sid)
            except User.DoesNotExist:
                return Response({'success': False, 'message': '学生不存在'}, status=404)
        s = SupervisionSerializer(obj, data=data, partial=True)
        if s.is_valid():
            kwargs = {}
            if student:
                kwargs['student'] = student
            try:
                s.save(**kwargs)
            except IntegrityError:
                return Response({'success': False, 'message': '该学生已有监管信息记录'}, status=400)
            return Response({'success': True, 'data': s.data})
        return Response({'success': False, 'message': s.errors}, status=400)
    elif request.method == 'DELETE':
        obj.delete()
        return Response({'success': True, 'message': '删除成功'})


# ─── 用户管理 ───────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def admin_user_list(request):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    if request.method == 'GET':
        qs = User.objects.all()
        filter_map = {
            'college': 'college__icontains',
            'major': 'major__icontains',
            'gender': 'gender',
            'batch': 'batch',
            'classNumber': 'class_number',
            'name': 'name__icontains',
            'studentId': 'student_id__icontains',
        }
        result = paginate(qs, request, search_fields=['name', 'student_id', 'college', 'major'], filter_map=filter_map)
        result['data'] = UserSerializer(result['data'], many=True).data
        # 如果请求 all_ids，返回所有 ID
        if request.query_params.get('all_ids'):
            result['ids'] = list(result['_filtered_qs'].values_list('student_id', flat=True))
        result.pop('_filtered_qs', None)
        return Response(result)
    # POST — 创建用户
    data = request.data.copy()
    # 前端用 camelCase，统一成模型字段名
    sid = data.get('student_id') or data.get('studentId', '')
    if not sid:
        return Response({'success': False, 'message': '学号不能为空'}, status=400)
    if User.objects.filter(student_id=sid).exists():
        return Response({'success': False, 'message': '该学号已存在'}, status=400)
    # 默认密码为学号（与批量导入一致），首次登录会被 needChangePassword 提示修改
    password = data.get('password') or sid
    user = User.objects.create_user(
        username=sid,
        student_id=sid,
        password=password,
        name=data.get('name', ''),
        gender=data.get('gender', ''),
        college=data.get('college', ''),
        major=data.get('major', ''),
        class_number=data.get('class_number') or data.get('classNumber', ''),
        batch=data.get('batch') or '无',
        role=data.get('role') or 'student',
    )
    return Response({'success': True, 'data': UserSerializer(user).data}, status=201)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_user_detail(request, student_id):
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    try:
        user = User.objects.get(student_id=student_id)
    except User.DoesNotExist:
        return Response({'success': False, 'message': '用户不存在'}, status=404)
    if request.method == 'GET':
        return Response(UserSerializer(user).data)
    elif request.method == 'PUT':
        s = UserSerializer(user, data=request.data, partial=True)
        if s.is_valid():
            s.save()
            return Response({'success': True, 'data': s.data})
        return Response({'success': False, 'message': s.errors}, status=400)
    elif request.method == 'DELETE':
        user.delete()
        return Response({'success': True, 'message': '删除成功'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_reset_password(request, student_id):
    """重置单个用户密码"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    try:
        user = User.objects.get(student_id=student_id)
        new_password = request.data.get('password', student_id)
        user.set_password(new_password)
        user.save()
        return Response({'success': True, 'message': '密码已重置'})
    except User.DoesNotExist:
        return Response({'success': False, 'message': '用户不存在'}, status=404)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_batch_reset_password(request):
    """批量重置密码"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    student_ids = request.data.get('student_ids', [])
    custom_password = request.data.get('password', None)
    count = 0
    errors = []
    for sid in student_ids:
        try:
            user = User.objects.get(student_id=sid)
            pwd = custom_password if custom_password else (user.student_id or '123456')
            user.set_password(pwd)
            user.save()
            count += 1
        except User.DoesNotExist:
            errors.append(f'用户 {sid} 不存在')
    msg = f'已重置 {count} 个用户密码'
    if errors:
        msg += '；' + '；'.join(errors)
    return Response({'success': True, 'message': msg})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_reset_all_passwords(request):
    """重置所有学生密码为学号"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    students = User.objects.exclude(role='admin')
    count = 0
    for user in students:
        pwd = user.student_id if user.student_id else '123456'
        user.set_password(pwd)
        user.save()
        count += 1
    return Response({'success': True, 'message': f'已重置 {count} 个学生密码'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_import_students(request):
    """导入学生（批量创建用户）- 上传 Excel 文件"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)

    from openpyxl import load_workbook
    import xlrd
    import re
    import os

    uploaded_file = request.data.get('file')
    batch = request.data.get('batch', '无')

    if not uploaded_file:
        return Response({'success': False, 'error': '请上传文件', 'imported': 0, 'skipped': 0, 'failed': 0})

    try:
        # 判断文件类型：.xls 用 xlrd，.xlsx 用 openpyxl
        filename = getattr(uploaded_file, 'name', '')
        is_xls = filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx')

        rows = []
        if is_xls:
            wb = xlrd.open_workbook(file_contents=uploaded_file.read())
            ws = wb.sheet_by_index(0)
            for r in range(1, ws.nrows):
                rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        else:
            wb = load_workbook(uploaded_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(list(row) if row else [])
    except Exception:
        return Response({'success': False, 'error': '无法解析文件，请确保是 .xlsx/.xls 格式', 'imported': 0, 'skipped': 0, 'failed': 0})

    imported = 0
    skipped = 0
    failed = 0
    errors = []

    for row in rows:
        if not row or not row[0]:
            continue
        sid = str(row[0]).strip() if row[0] is not None else ''
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        gender = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        raw_class = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''

        if not sid:
            continue

        if User.objects.filter(student_id=sid).exists():
            skipped += 1
            continue

        # 从班级中提取专业和班级序号
        # 例如: 2023生物医学工程（创新班）01 → 专业=生物医学工程（创新班）, 班级序号=01
        major = ''
        class_number = ''
        if raw_class:
            m = re.search(r'\d+\s*(.+?)\s*(\d{2})$', raw_class)
            if m:
                major = m.group(1).strip()
                class_number = m.group(2).strip()

        try:
            User.objects.create_user(
                username=sid,
                student_id=sid,
                password=sid,
                name=name,
                gender=gender,
                college='医学部',
                major=major,
                class_number=class_number,
                batch=batch,
            )
            imported += 1
        except Exception as e:
            failed += 1
            errors.append(str(sid) + ': ' + str(e))

    result = {
        'success': True,
        'imported': imported,
        'skipped': skipped,
        'failed': failed,
        'message': f'成功导入 {imported} 个学生',
    }
    if errors:
        result['errors'] = errors
    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_batch_delete(request, entity_type):
    """批量删除"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    ids = request.data.get('ids', [])
    model_map = {
        'internship': Internship,
        'registration': Registration,
        'supervision': Supervision,
        'users': User,
    }
    model = model_map.get(entity_type)
    if not model:
        return Response({'success': False, 'message': '未知类型'}, status=400)
    # User 用 student_id 作为主键
    if model == User:
        deleted, _ = model.objects.filter(student_id__in=ids).delete()
    else:
        # Internship/Registration/Supervision 用 student_id 作为主键（OneToOne PK）
        deleted, _ = model.objects.filter(student_id__in=ids).delete()
    return Response({'success': True, 'message': f'已删除 {deleted} 条记录'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_batch_export(request, entity_type):
    """批量导出为 ZIP（实习报告/登记表）或汇总 Excel（监管信息）"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)

    student_ids = request.data.get('studentIds', [])
    if not student_ids:
        return Response({'success': False, 'message': '没有要导出的数据'}, status=400)

    from django.http import FileResponse

    try:
        if entity_type == 'internship':
            rows = Internship.objects.select_related('student').filter(student_id__in=student_ids)
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for obj in rows:
                    data = obj.to_export_dict()
                    fs = generate_report(data)
                    zf.writestr(f"{data['studentId']}_{data['name']}_实习报告.docx", fs.getvalue())
            buf.seek(0)
            count = rows.count()
            return FileResponse(buf, as_attachment=True,
                filename=f"{datetime.now():%m-%d-%M}_批量导出{count}条_实习报告.zip")

        elif entity_type == 'registration':
            rows = Registration.objects.select_related('student').filter(student_id__in=student_ids)
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for obj in rows:
                    data = obj.to_export_dict()
                    fs = generate_registration_excel(data)
                    zf.writestr(f"{data['studentId']}_{data['name']}_实习情况登记表.xlsx", fs.getvalue())
            buf.seek(0)
            count = rows.count()
            return FileResponse(buf, as_attachment=True,
                filename=f"{datetime.now():%m-%d-%M}_批量导出{count}条_实习情况登记表.zip")

        elif entity_type == 'supervision':
            rows = Supervision.objects.select_related('student').filter(student_id__in=student_ids)
            rows_data = [obj.to_export_dict() for obj in rows]
            fs = generate_supervision_batch_excel(rows_data)
            count = len(rows_data)
            return FileResponse(fs, as_attachment=True,
                filename=f"{datetime.now():%m-%d-%M}_批量导出{count}条_实习监管信息表.xlsx")

        return Response({'success': False, 'message': '未知类型'}, status=400)
    except Exception as e:
        return Response({'success': False, 'message': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_export_detail(request, entity_type, student_id):
    """导出单个记录"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)

    from django.http import FileResponse
    from internship_system.docgen import generate_supervision_excel

    if entity_type == 'supervision':
        try:
            obj = Supervision.objects.select_related('student').get(student_id=student_id)
        except Supervision.DoesNotExist:
            return Response({'success': False, 'message': '记录不存在'}, status=404)
        s = obj.student
        data = {
            'studentId': s.student_id, 'name': obj.name,
            'enrollmentYear': obj.enrollment_year, 'college': obj.college,
            'classNumber': obj.class_number, 'courseName': obj.course_name,
            'courseCode': obj.course_code, 'credits': obj.credits,
            'internshipType': obj.internship_type, 'internshipOrgForm': obj.internship_org_form,
            'internshipMethod': obj.internship_method, 'academicYear': obj.academic_year,
            'teacher': obj.teacher, 'company': obj.company,
            'companyCreditCode': obj.company_credit_code,
            'internshipAreaCode': obj.internship_area_code,
            'internshipAddress': obj.internship_address,
            'hasLiabilityInsurance': obj.has_liability_insurance,
            'hasAccidentInsurance': obj.has_accident_insurance,
            'hasSafetyTraining': obj.has_safety_training,
            'hasTripartiteAgreement': obj.has_tripartite_agreement,
            'companyAcquisitionMethod': obj.company_acquisition_method,
            'hasEmergencyPlan': obj.has_emergency_plan,
            'isSchoolBase': obj.is_school_base, 'isDigitalBase': obj.is_digital_base,
            'isOverseasBase': obj.is_overseas_base,
            'startDate': str(obj.start_date or ''), 'endDate': str(obj.end_date or ''),
            'totalDays': str(obj.total_days),
            'internshipPosition': obj.internship_position,
            'internshipSalary': obj.internship_salary,
            'companySupervisor': obj.company_supervisor,
        }
        fs = generate_supervision_excel(data)
        return FileResponse(fs, as_attachment=True,
                            filename=f"{s.student_id}_{obj.name}_实习监管信息表.xlsx")

    return Response({'success': False, 'message': '该类型不支持导出'}, status=400)


DEFAULT_SETTINGS = {
    'courseName': '专业实习',
    'courseCode': 'CSE450',
    'credits': '4（4）',
    'internshipType': '专业实习',
    'internshipOrgForm': '集中实习',
    'internshipMethod': '现场实习',
    'academicYear': '2024-2025学年',
    'isSchoolBase': '否',
    'isDigitalBase': '否',
    'isOverseasBase': '否',
}


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def admin_settings(request):
    """获取/保存系统配置（管理员用）"""
    if not request.user.is_admin:
        return Response({'success': False, 'message': '权限不足'}, status=403)
    if request.method == 'GET':
        # 从数据库读取，没有则用默认值
        result = {}
        for key in DEFAULT_SETTINGS:
            try:
                obj = SystemSetting.objects.get(key=key)
                result[key] = obj.value
            except SystemSetting.DoesNotExist:
                result[key] = DEFAULT_SETTINGS[key]
        return Response(result)
    # PUT — 保存到数据库
    for key, value in request.data.items():
        if key in DEFAULT_SETTINGS:
            SystemSetting.objects.update_or_create(
                key=key, defaults={'value': str(value)}
            )
    return Response({'success': True, 'message': '配置已保存'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def public_settings(request):
    """获取系统配置（所有登录用户可用，只读）"""
    result = {}
    for key in DEFAULT_SETTINGS:
        try:
            obj = SystemSetting.objects.get(key=key)
            result[key] = obj.value
        except SystemSetting.DoesNotExist:
            result[key] = DEFAULT_SETTINGS[key]
    return Response(result)
